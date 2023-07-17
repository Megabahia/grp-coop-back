import io

from django.core.files.uploadedfile import InMemoryUploadedFile
from .models import PagoEmpleados
from .serializers import (
    PagoEmpleadosSerializer
)
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
import json
# excel
import openpyxl
# Utils
from ...utils import utils
# Import PDF
from fpdf import FPDF
# Enviar Correo
from apps.config.util import sendEmail
# ObjectId
from bson import ObjectId
# logs
from apps.CENTRAL.central_logs.methods import createLog, datosTipoLog, datosProductosMDP

# declaracion variables log
datosAux = datosProductosMDP()
datosTipoLogAux = datosTipoLog()
# asignacion datos modulo
logModulo = datosAux['modulo']
logApi = datosAux['api']
# asignacion tipo de datos
logTransaccion = datosTipoLogAux['transaccion']
logExcepcion = datosTipoLogAux['excepcion']


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_pagosEmpleados(request):
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            first = True  # si tiene encabezado
            uploaded_file = request.FILES['documento']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(uploaded_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Pagos"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if len(dato) == 10:
                    resultadoInsertar = insertarDato_PagoEmpleado(dato, request.data['user_id'])
                    if resultadoInsertar != 'Dato insertado correctamente':
                        contInvalidos += 1
                        errores.append({"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        return Response(result, status=status.HTTP_201_CREATED)

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def pagoEmpleados_update(request, pk):
    request.POST._mutable = True
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'update/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'ESCRIBIR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            logModel['dataEnviada'] = str(request.data)
            query = PagoEmpleados.objects.filter(pk=ObjectId(pk), state=1).first()
        except PagoEmpleados.DoesNotExist:
            errorNoExiste = {'error': 'No existe'}
            createLog(logModel, errorNoExiste, logExcepcion)
            return Response(status=status.HTTP_404_NOT_FOUND)
        if request.method == 'POST':
            now = timezone.localtime(timezone.now())
            request.data['updated_at'] = str(now)
            if 'created_at' in request.data:
                request.data.pop('created_at')

            if 'claveFirma' in request.data:
                if request.data != '':
                    print('llega')
                    datau, datas = firmar(request)
                    archivo_pdf_para_enviar_al_cliente = io.BytesIO()
                    archivo_pdf_para_enviar_al_cliente.write(datau)
                    archivo_pdf_para_enviar_al_cliente.write(datas)
                    archivo_pdf_para_enviar_al_cliente.seek(0)

                    request.data['archivoFirmado'] = InMemoryUploadedFile(archivo_pdf_para_enviar_al_cliente,
                                                                          'media',
                                                                          'documentoFirmado.pdf',
                                                                          'application/pdf',
                                                                          archivo_pdf_para_enviar_al_cliente.getbuffer().nbytes,
                                                                          None
                                                                          )
                    request.data['fechaFirma'] = str(now)

            serializer = PagoEmpleadosSerializer(query, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                print(serializer.data['empresa'])

                if serializer.data['estado'] == 'Negado':
                    registro = serializer.data
                    envioCorreoNegado(registro['correo'], registro['nombresCompletos'], registro['observacion'])
                if serializer.data['estado'] == 'Aprobado':
                    registro = serializer.data
                    envioCorreoAprobado(registro['empresa']['correo'], registro['nombresCompletos'],
                                        registro['montoPagar'], registro['montoDisponible'])
                    nombrePyme = registro['empresa']['comercial']
                    nombreReresentanteLegal = registro['empresa']['reprsentante']
                    envioCorreoTranserencia(registro['correo'], registro['montoPagar'], registro['nombresCompletos'],
                                            nombreReresentanteLegal, nombrePyme, registro['mesPago'])

                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def pagoEmpleados_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}

            if "user_id" in request.data and request.data["user_id"]:
                filters['user_id'] = str(request.data["user_id"])

            # Serializar los datos
            query = PagoEmpleados.objects.filter(**filters).order_by('-created_at')
            serializer = PagoEmpleadosSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


from cryptography.hazmat import backends
from cryptography.hazmat.primitives.serialization import pkcs12
from endesive.pdf import cms


def firmar(request):
    usuario = json.loads(request.data['usuarioEmpresa'])
    certificado = request.data['certificado']
    pdf = request.data['archivoFirmado']
    contrasenia = request.data['claveFirma']
    date = timezone_now = timezone.localtime(timezone.now())
    date = date.strftime("D:%Y%m%d%H%M%S+00'00'")
    dct = {
        "aligned": 0,
        "sigflags": 3,
        "sigflagsft": 132,
        "sigpage": 0,
        "sigbutton": True,
        "sigfield": "Signature1",
        "auto_sigfield": True,
        "sigandcertify": True,
        "signaturebox": (470, 840, 570, 640),
        "signature": usuario['nombresCompleto'],
        # "signature_img": "signature_test.png",
        "contact": usuario['email'],
        "location": "Ubicación",
        "signingdate": date,
        "reason": "Pago a empleado",
        "password": contrasenia,
    }
    # with open("cert.p12", "rb") as fp:
    p12 = pkcs12.load_key_and_certificates(
        certificado.read(), contrasenia.encode("ascii"), backends.default_backend()
    )

    # datau = open(fname, "rb").read()
    datau = pdf.read()
    datas = cms.sign(datau, dct, p12[0], p12[1], p12[2], "sha256")
    return datau, datas
    # return Response('new_serializer_data', status=status.HTTP_200_OK)


def enviarNegadoPago(email, monto):
    subject, from_email, to = 'RAZÓN POR LA QUE SE NIEGA EL PAGO A PROVEEDORES', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
        PAGO A PROVEEDORES - CRÉDITO PAGOS

        Lo sentimos!!

        La transferencia por ${monto} DE LA FACTURA A PAGAR ha sido rechazada. 
        Por favor revise sus fondos e intente de nuevo.

        Si cree que es un error, contáctese con su agente a través de https://walink.co/b5e9c0

        Atentamente,

        Global RedPyme – Crédito Pagos
    """
    html_content = f"""
                <html>
                    <body>
                        <h1>PAGO A PROVEEDORES - CRÉDITO PAGOS</h1>
                        <br>
                        <h3><b>Lo sentimos!!</b></h3>
                        <br>
                        <p>La transferencia por ${monto} DE LA FACTURA A PAGAR ha sido rechazada. 
                        Por favor revise sus fondos e intente de nuevo.</p>
                        <br>
                        <br>
                        <p>Si cree que es un error, contáctese con su agente a través de https://walink.co/b5e9c0</p>
                        <br>
                        Atentamente,
                        <br>
                        Global RedPyme – Crédito Pagos
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)


def enviarProcesandoPago(email, monto):
    subject, from_email, to = 'Transferencia exitosa', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
        PAGO A PROVEEDORES - CRÉDITO PAGOS

        FELICIDADES!!

        La transferencia por ${monto} ha sido realizada con éxito, 
        adjuntamos el comprobante del pago realizado. 
        En 24 horas será acreditado a la cuenta destino.

        Atentamente,

        Global RedPyme – Crédito Pagos
    """
    html_content = f"""
                <html>
                    <body>
                        <h1>PAGO A PROVEEDORES - CRÉDITO PAGOS</h1>
                        <br>
                        <h3><b>FELICIDADES!!</b></h3>
                        <br>
                        <p>La transferencia por ${monto} ha sido realizada con éxito, 
                        adjuntamos el comprobante del pago realizado. 
                        En 24 horas será acreditado a la cuenta destino.</p>
                        <br>
                        <br>
                        Atentamente,
                        <br>
                        Global RedPyme – Crédito Pagos
                        <br>
                    </body>
                </html>
                """
    sendEmail(subject, txt_content, from_email, to, html_content)


def generarPDF(datos):
    # save FPDF() class into a
    # variable pdf
    pdf = FPDF()

    # Add a page
    pdf.add_page()

    # set style and size of font
    # that you want in the pdf
    pdf.set_font("Arial", size=15)

    # create a cell
    pdf.cell(200, 10, txt="GeeksforGeeks",
             ln=1, align='C')

    # add another cell
    pdf.cell(200, 10, txt="A Computer Science portal for geeks.",
             ln=2, align='C')

    return pdf


def insertarDato_PagoEmpleado(dato, user_id):
    try:
        if ('None' in dato):
            return 'Tiene campos vacios'
        if (not utils.__validar_ced_ruc(dato[5], 0)):
            return 'Cedula incorrecta'
        if (not utils.isValidEmail(dato[9])):
            return 'Email incorrecto'
        if (not utils.isValidTelefono(dato[8])):
            return 'Celular incorrecto'
        if (not utils.isValidTelefono(dato[8])):
            return 'Whatsapp incorrecto'
        data = {}
        data['nombresCompletos'] = f"{dato[6]} {dato[7]}"
        data['cedula'] = dato[5]
        data['celular'] = dato[8]
        data['correo'] = dato[9]
        data['montoPagar'] = dato[3]
        data['codigoEmpleado'] = dato[4]
        data['mesPago'] = dato[1]
        data['anio'] = dato[2]
        data['estado'] = ''
        data['user_id'] = user_id
        data['state'] = 1
        # inserto el dato con los campos requeridos
        PagoEmpleados.objects.update_or_create(**data)
        return 'Dato insertado correctamente'
    except Exception as e:
        return str(e)


def envioCorreoNegado(email, nombresCompletosEmpleado, observacion):
    subject, from_email, to = 'Solicitud de Pago a Empleado NEGADA', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
            Solicitud de Pago a Empleado NEGADA

            Lo sentimos!!

            El pago a {nombresCompletosEmpleado} ha sido NEGADO debido a {observacion}

            Si requiere ayuda, contáctese con un asesor a través del siguiente enlace: https://wa.link/nczlei

            Atentamente,

            Global RedPyme – Crédito Pagos
        """
    html_content = f"""
        <html>
            <body>
                <h1>Solicitud de Pago a Empleado NEGADA</h1>
                <br>
                <h3><b>Lo sentimos!!</b></h3>
                <br>
                <p>El pago a {nombresCompletosEmpleado} ha sido NEGADO debido a {observacion}</p>
                <br>
                <br>
                <p>Si requiere ayuda, contáctese con un asesor a través del siguiente enlace: https://wa.link/nczlei</p>
                <br>
                Atentamente,
                <br>
                Equipo Global Redpyme – Crédito Pagos
                <br>
            </body>
        </html>
    """
    sendEmail(subject, txt_content, from_email, to, html_content)


def envioCorreoAprobado(email, nombresCompletosEmpleado, monto, montoDisponible):
    subject, from_email, to = 'Solicitud de Pago a Empleado APROBADA', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
            Solicitud de Pago a Empleado APROBADA

            LISTO,

            El pago a {nombresCompletosEmpleado} un monto de {monto} a sido APROBADO. Su Monto Disponible de Línea de Crédito es: {montoDisponible}

            Atentamente,

            Equipo Global Redpyme – Crédito Pagos
        """
    html_content = f"""
        <html>
            <body>
                <h1>Solicitud de Pago a Empleado APROBADA</h1>
                <br>
                <h3><b>LISTO,</b></h3>
                <br>
                <p>El pago a {nombresCompletosEmpleado} un monto de {monto} a sido APROBADO. 
                Su Monto Disponible de Línea de Crédito es: {montoDisponible}</p>
                <br>
                <br>
                Atentamente,
                <br>
                Equipo Global Redpyme – Crédito Pagos
                <br>
            </body>
        </html>
    """
    sendEmail(subject, txt_content, from_email, to, html_content)


def envioCorreoTranserencia(email, monto, nombresCompletosEmpleado, nombreRepresentanteLegal, nombrePyme, mesPago):
    subject, from_email, to = 'TRANSFERENCIA RECIBIDA', "08d77fe1da-d09822@inbox.mailtrap.io", \
                              email
    txt_content = f"""
            TRANSFERENCIA RECIBIDA

            Usted ha recibido una transferencia por parte de {nombreRepresentanteLegal},

            Estimado/o {nombresCompletosEmpleado} usted ha recibido una transferencia en su cuenta bancaria por un 
            monto de {monto} por sus labores en la empresa {nombrePyme} respecto al mes de {mesPago}

            Atentamente,

            Equipo Global Redpyme – Crédito Pagos
        """
    html_content = f"""
        <html>
            <body>
                <h1>TRANSFERENCIA RECIBIDA</h1>
                <br>
                <h3><b>Usted ha recibido una transferencia por parte de {nombreRepresentanteLegal},</b></h3>
                <br>
                <p>Estimado/o {nombresCompletosEmpleado} usted ha recibido una transferencia en su cuenta bancaria por
                 un monto de {monto} por sus labores en la empresa {nombrePyme}  respecto al mes de {mesPago}
                 </p>
                <br>
                <br>
                Atentamente,
                <br>
                Equipo Global Redpyme – Crédito Pagos
                <br>
            </body>
        </html>
    """
    sendEmail(subject, txt_content, from_email, to, html_content)
